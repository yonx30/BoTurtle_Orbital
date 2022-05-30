import os
import logging
import openpyxl as op

from telegram import (
  Update, 
  ForceReply,
  InlineQuery,
  InlineKeyboardButton,
  InlineKeyboardMarkup,
) #upm package(python-telegram-bot) 
from telegram.ext import (
  Updater, 
  CommandHandler, 
  MessageHandler, 
  ConversationHandler, 
  InlineQueryHandler,
  CallbackQueryHandler, 
  Filters, 
  CallbackContext,
) #upm package(python-telegram-bot)

# Enable logging
logging.basicConfig(
  filename="log.txt",
  filemode='w',
  format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
  level=logging.INFO
)

logger = logging.getLogger(__name__)

token = '5193620467:AAE0uvyuVGDxvDoqIGG91XKF6HFx5b_Lz18'
#token = os.environ['CADit_token']


# Stages
REGISTER, CLIENT_INFO, COMPLETE = range(3)
# Common callback data
SELLER, CLIENT, NEXT, END, RESTART = range(5)
# Material callback data
OPTION1, OPTION2, OPTION3 = range(3)
# Client stages
STL, MATERIAL, COLOR, INFILL, ORIENTATION, CHECK = range(6)


def get_excel_val(sheet, col, row):
  return sheet["%s%s" % (col, row)].value

def write_excel_val(sheet, col, row, input, font=None, fill=None, *args):
  sheet["%s%s" % (col, row)] = input
  if font:
      if "bold" in font:
          sheet["%s%s" % (col, row)].font = op.styles.Font(bold=True)
  if fill:
      sheet["%s%s" % (col, row)].fill = op.styles.PatternFill(patternType = "darkDown", start_color=fill)


# Define a few command handlers. These usually take the two arguments update and context.
def start(update: Update, _: CallbackContext) -> int:
  global userid, user_handle
  userid = update.message.from_user
  user_handle = userid["username"]

  """Send a message when the command /start is issued."""
  user = update.effective_user
  logger.info("User %s started the conversation.", user.first_name)
  
  keyboard = [
        [
            InlineKeyboardButton("Seller", callback_data=str(SELLER)),
            InlineKeyboardButton("Client", callback_data=str(CLIENT)),
        ]
    ]
  reply_markup = InlineKeyboardMarkup(keyboard)
  # Send message with text and appended Inlinekeyboard
  update.message.reply_text(fr"Hi {user.first_name}! Do you want to register as a seller or client?", reply_markup=reply_markup)

  return REGISTER

def access_excel():
  global workbook_path, wb, vendor_sheet_name, vendor_sheet, client_sheet_name, client_sheet
  workbook_path = os.getcwd() + "/" # Gets the directory the Python script is in'''
  try:
    wb = op.load_workbook(filename = workbook_path + "User list.xlsx", read_only=False, data_only=True) # Opens in edit mode to add in extra sheets, and only reads the values & not formulaes
  except FileNotFoundError:
    print("Error: File not found")
  else:
    sheet_list = wb.sheetnames
    vendor_sheet_name = sheet_list[0]
    vendor_sheet = wb[vendor_sheet_name]
    client_sheet_name = sheet_list[1]
    client_sheet = wb[client_sheet_name]
    

def restart(update: Update, _: CallbackContext) -> int:

  """Restart the ordering process for the client."""
  query = update.callback_query
  query.answer()
  keyboard = [
        [
            InlineKeyboardButton("Seller", callback_data=str(SELLER)),
            InlineKeyboardButton("Client", callback_data=str(CLIENT)),
        ]
    ]
  reply_markup = InlineKeyboardMarkup(keyboard)
  # Send message with text and appended Inlinekeyboard
  update.message.reply_text("Hi! Do you want to register as a seller or client?", reply_markup=reply_markup)

  return REGISTER

def help_command(update: Update, _: CallbackContext) -> None:
  """Send a message when the command /help is issued."""
  update.message.reply_text('Help!')

def downloader(update, context):
    context.bot.get_file(update.message.document).download()

    # writing to a custom file
    with open("custom/file.doc", 'wb') as f:
        context.bot.get_file(update.message.document).download(out=f)

def echo(update: Update, _: CallbackContext) -> None:
  """Echo the user message."""
  update.message.reply_text(f"you said {update.message.text}")


def handle_msg(update, context):
    text = str(update.message.text).lower()
    response = response_msg(text)
    update.message.reply_text("Processing message... \n %s"%(response))

def handle_photo(update, context):
    try:
        photo = update.message.photo
    except:
        update.message.reply_text("That's not a photo!")
    else:
        update.message.reply_text("Here's your photo back!")
        update.message.reply_photo(photo)

def error(update, context):
    print("Update %s cause error %s"%(update, context.error))


def caps(update: Update, context: CallbackContext):
    text_caps = ' '.join(context.args).upper()
    context.bot.send_message(chat_id=update.effective_chat.id, text=text_caps)
    #context.bot.send_message(text_caps)
  



def seller(update: Update, context: CallbackContext) -> int:
  for row in range(2, 100):
    print(get_excel_val(vendor_sheet, "B", row))
    if get_excel_val(vendor_sheet, "B", row) == user_handle:
      break
    elif get_excel_val(vendor_sheet, "B", row) == None:
      write_excel_val(vendor_sheet, "A", row, row-1)
      col = 66
      for i in range(5):
        print(get_excel_val(vendor_sheet, chr(col), 1))
        print(userid[get_excel_val(vendor_sheet, chr(col), 1)])
        write_excel_val(vendor_sheet, chr(col), row, userid[get_excel_val(vendor_sheet, chr(col), 1)])
        col += 1
      wb.save(filename= workbook_path + "User List.xlsx")
      break
  """TO BE EDITED TO ASK FOR ANY NECESSARY"""
  query = update.callback_query
  querydata = query.data
  context.user_data["Q1"] = querydata
  query.answer()
  keyboard = [
    [
      InlineKeyboardButton("FDM", callback_data=str(MATERIAL)),
      InlineKeyboardButton("Resin", callback_data=str(MATERIAL)),
    ],
    [
      InlineKeyboardButton("SLA", callback_data=str(MATERIAL)),
    ]
  ]
  reply_markup = InlineKeyboardMarkup(keyboard)
  query.edit_message_text(
    text="What kind of machine do you use?", reply_markup=reply_markup
  )
  return CLIENT_INFO
  

def client(update: Update, context: CallbackContext) -> int:
  for row in range(2, 100):
    print(get_excel_val(client_sheet, "B", row))
    if get_excel_val(client_sheet, "B", row) == user_handle:
      break
    elif get_excel_val(client_sheet, "B", row) == None:
      write_excel_val(client_sheet, "A", row, row-1)
      col = 66
      for i in range(5):
        print(get_excel_val(client_sheet, chr(col), 1))
        print(userid[get_excel_val(client_sheet, chr(col), 1)])
        write_excel_val(client_sheet, chr(col), row, userid[get_excel_val(client_sheet, chr(col), 1)])
        col += 1
      wb.save(filename= workbook_path + "User List.xlsx")
      break
  """TO BE EDITED TO ASK FOR ANY NECESSARY"""
  query = update.callback_query
  querydata = query.data
  context.user_data["Q1"] = querydata
  query.answer()
  keyboard = [
    [
      InlineKeyboardButton("Submit Order", callback_data=str(MATERIAL)),
      InlineKeyboardButton("To be added", callback_data=str(MATERIAL)),
    ],
    [
      InlineKeyboardButton("To be added", callback_data=str(MATERIAL)),
    ]
  ]
  reply_markup = InlineKeyboardMarkup(keyboard)
  query.edit_message_text(
    text="What would you like to do?", reply_markup=reply_markup
  )
  return CLIENT_INFO

def client_stl(update: Update, context: CallbackContext) -> int:
  context.bot.get_file(update.message.document).download()

  '''#writing to a custom file
  with open("custom file.stl", 'wb') as f:
    context.bot.get_file(update.message.document).download(out=f)'''
  
  

def client_material(update: Update, context: CallbackContext) -> int:
  """Material choices for client"""
  query = update.callback_query
  querydata = query.data
  context.user_data["Material"] = querydata
  query.answer()
  keyboard = [
    [
      InlineKeyboardButton("Option 1", callback_data=str(COLOR)),
      InlineKeyboardButton("Option 2", callback_data=str(COLOR)),
    ],
    [
      InlineKeyboardButton("Option 3", callback_data=str(COLOR)),
    ]
  ]
  reply_markup = InlineKeyboardMarkup(keyboard)
  query.edit_message_text(
    text="Choose the desired material for your 3D print", reply_markup=reply_markup
  )
  return CLIENT_INFO
  
def client_color(update: Update, context: CallbackContext) -> int:
  """Color choices for client"""
  query = update.callback_query
  querydata = query.data
  context.user_data["Color"] = querydata
  query.answer()
  keyboard = [
    [
      InlineKeyboardButton("Option 1", callback_data=str(INFILL)),
      InlineKeyboardButton("Option 2", callback_data=str(INFILL)),
    ],
    [
      InlineKeyboardButton("Option 3", callback_data=str(INFILL)),
    ]
  ]
  reply_markup = InlineKeyboardMarkup(keyboard)
  query.edit_message_text(
    text="Choose the desired color for your 3D print", reply_markup=reply_markup
  )
  return CLIENT_INFO

def client_infill(update: Update, context: CallbackContext) -> int:
  """Infill choices for client"""
  query = update.callback_query
  querydata = query.data
  context.user_data["Infill"] = querydata
  query.answer()
  keyboard = [
    [
      InlineKeyboardButton("Option 1", callback_data=str(ORIENTATION)),
      InlineKeyboardButton("Option 2", callback_data=str(ORIENTATION)),
    ],
    [
      InlineKeyboardButton("Option 3", callback_data=str(ORIENTATION)),
    ]
  ]
  reply_markup = InlineKeyboardMarkup(keyboard)
  query.edit_message_text(
    text="Choose the desired infill for your 3D print", reply_markup=reply_markup
  )
  return CLIENT_INFO

def client_orientation(update: Update, context: CallbackContext) -> int:
  """Orientation choices for client"""
  query = update.callback_query
  querydata = query.data
  context.user_data["Orientation"] = querydata
  query.answer()
  
  keyboard = [
    [
      InlineKeyboardButton("Option 1", callback_data=str(CHECK)),
      InlineKeyboardButton("Option 2", callback_data=str(CHECK)),
    ],
    [
      InlineKeyboardButton("Option 3", callback_data=str(CHECK)),
    ]
  ]
  reply_markup = InlineKeyboardMarkup(keyboard)
  query.edit_message_text(
    text="Choose the desired orientation for your 3D print", reply_markup=reply_markup
  )
  return CLIENT_INFO

def check_order(update: Update, context: CallbackContext) -> int:
  """Orientation choices for client"""
  query = update.callback_query
  query.answer()
  
  keyboard = [
    [
      InlineKeyboardButton("Restart", callback_data=str(RESTART)),
      InlineKeyboardButton("Confirm Order", callback_data=str(END)),
    ],
  ]
  reply_markup = InlineKeyboardMarkup(keyboard)
  query.edit_message_text(
    text="Confirm your order?", reply_markup=reply_markup
  )
  return COMPLETE

def end(update: Update, context: CallbackContext) -> int:
    """Returns `ConversationHandler.END`, which tells the
    ConversationHandler that the conversation is over.
    """
    query = update.callback_query
    query.answer()
    query.edit_message_text(text="See you next time!")
    return ConversationHandler.END

def main() -> None:
  """Start the bot."""
  # Create the Updater and pass it your bot's token.
  updater = Updater(token)

  # Get the dispatcher to register handlers
  dp = updater.dispatcher

  conv_handler = ConversationHandler(
    entry_points=[CommandHandler('start', start)],
    states={
      REGISTER: [
        CallbackQueryHandler(seller, pattern='^' + str(SELLER) + '$'),
        CallbackQueryHandler(client, pattern='^' + str(CLIENT) + '$'),
      ],
      CLIENT_INFO: [        
        CallbackQueryHandler(client_material, pattern='^' + str(MATERIAL) + '$'),
        CallbackQueryHandler(client_color, pattern='^' + str(COLOR) + '$'),
        CallbackQueryHandler(client_infill, pattern='^' + str(INFILL) + '$'),
        CallbackQueryHandler(client_orientation, pattern='^' + str(ORIENTATION) + '$'),
        CallbackQueryHandler(check_order, pattern='^' + str(CHECK) + '$'),
      ],
      COMPLETE: [
        CallbackQueryHandler(restart, pattern='^' + str(RESTART) + '$'),
        CallbackQueryHandler(end, pattern='^' + str(END) + '$'),
      ],
    },
    fallbacks=[CommandHandler('start', start)],
  )

  # Add ConversationHandler to dispatcher that will be used for handling updates
  dp.add_handler(conv_handler)
  dp.add_handler(MessageHandler(Filters.document, downloader))

  # on different commands - answer in Telegram
  # dp.add_handler(CommandHandler("start", start))
  # dp.add_handler(CommandHandler("help", help_command))

  # on non command i.e message - echo the message on Telegram
  # dp.add_handler(MessageHandler(Filters.text, echo))
  
  # dp.add_handler(CommandHandler('buttons', button_cmd))
  # dp.add_handler(MessageHandler(Filters.text & (~Filters.command), handle_msg)) #Adds message handler for text using handle_msg function
  
  # dp.add_handler(MessageHandler(Filters.photo, handle_photo))  
  # dp.add_handler(CommandHandler('caps', caps))
  # dp.add_handler(InlineQueryHandler(inline_caps))
  
  # Start the Bot
  updater.start_polling()

  # Run the bot until you press Ctrl-C or the process receives SIGINT,
  # SIGTERM or SIGABRT. This should be used most of the time, since
  # start_polling() is non-blocking and will stop the bot gracefully.
  updater.idle()


if __name__ == '__main__':
  access_excel()
  main()